from openpyxl import load_workbook
import time
import csv
import os
from ftplib import FTP


# ftp 获取ZTE过去三天日报
def ftp_getZteFile():
    ftp = FTP()
    # ftp.set_debuglevel(2)
    ftp.connect("124.75.32.237", 21)
    ftp.login("zxin10", "Sos10+Pad6")
    print(ftp.getwelcome())
    ftp.cwd("4ktongji")
    print(ftp.dir())
    localdir = "D:\日报\\"
    if not os.path.exists(localdir):
        os.makedirs(localdir)
    listf = ftp.nlst()
    for file in listf:
        local = os.path.join(localdir, file)
        print("下载", ftp.nlst(file))
        file_handler = open(local, 'wb').write
        ftp.retrbinary('RETR ' + file, file_handler)
    ftp.quit()

# 获取日，用于读取过去三天的日报
def getdate(i,day,m,month):
    inday = int(day) - i
    inday2 = int(day) - i + 1
    if (inday < 1):
        remon = int(m) - 1
        inday = month[remon] + inday
    elif (inday < 10):
        inday = "0{}".format(inday)
    if (inday2 < 1):
        remon = int(m) - 1
        inday2 = month[remon] + inday2
    elif (inday2 < 10):
        inday2 = "0{}".format(inday2)
    return inday,inday2

def getribao(m,day,bz_list,month, wbmb , bz, num):
    # wbmb = load_workbook("D:\日报\烽火并发日报模板.xlsx")
    wsmb = wbmb.active
    for i in range(1, num):
        inday1, inday2 = getdate(i, day, m, month)
        if bz == 0:
            inday1 = inday2
            onedayUrl = "D:\日报\ottnodecpservicestatistics-2021{}{}.csv"
            en = "gbk"
        else:
            onedayUrl = "D:\日报\ss_2021-{}-{}.csv"
            en = "utf-8"
        file = onedayUrl.format(m, inday1)
        print(file)
        csvF = open(file, 'r', encoding= en)
        read = csv.reader(csvF)  
        index = bz_list[i - 1]
        wsmb.cell(index, 1).value = "2021/{}/{}".format(m, inday2)
        index = index + 1
        for line in read:
            # print(line)
            k = 1
            for v in line:
                wsmb.cell(index, k).value = v
                k = k + 1
            index = index + 1


def main():
    d = input("请输入数字 1：生成一天 2：生成三天")
    num = 4
    if int(d) == 1:
        num = 2
    m = time.strftime("%m", time.localtime())
    day = time.strftime("%d", time.localtime())
    fhbz_list = [25, 13, 1]
    month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # month list
    startday = int(day) - 2

    wbmb = load_workbook("D:\日报\烽火并发日报模板.xlsx")
    getribao(m, day, fhbz_list, month, wbmb, 1, num)  # bz = 1 代表fh日报
    if num == 4:
        wbmb.save("D:\日报\烽火并发日报2021-{}-{}到{}.xlsx".format(m, startday, day))
    else:
        wbmb.save("D:\日报\烽火并发日报2021-{}-{}.xlsx".format(m, day))

    ftp_getZteFile()  # 获取中兴日报文件
    ztebz_list = [17, 9, 1]
    wbmb2 = load_workbook("D:\日报\中兴并发日报模板2.xlsx")
    getribao(m, day, ztebz_list, month, wbmb2, 0, num)  # bz = 0 代表zte日报
    if num == 4:
        wbmb2.save("D:\日报\中兴并发日报2021-{}-{}到{}.xlsx".format(m, startday, day))
    else:
        wbmb2.save("D:\日报\中兴并发日报2021-{}-{}.xlsx".format(m, startday, day))




if __name__ == '__main__':
   main()