from openpyxl import Workbook
from openpyxl import load_workbook
import time
import csv
import pandas as pd
import random
from copy import copy, deepcopy

# csvf= pd.read_csv('D:\日报\中兴并发日报模板.csv' ,encoding="gbk", header=None)
# print(csvf)
# csvf.to_csv('D:\日报\中兴并发日报模板test3.csv')

csvf = open("D:\日报\中兴并发日报模板test.csv", 'w', encoding='gbk',newline='')
csvtar = open("D:\日报\中兴并发日报模板.csv",encoding='gbk')
tarread = csv.reader(csvtar)
csfw = csv.writer(csvf)
day = '2021/4/10'
csfw.writerow(day,' ')
csfw.writerows(tarread)
# for r in tarread:
#     print(r)
#     csfw.writerow(r)



# # 烽火三天日报测试脚本
# m = time.strftime("%m",time.localtime())
# day = time.strftime("%d",time.localtime())
# wbmb = load_workbook("D:\日报\烽火并发日报模板.xlsx")
# wsmb = wbmb.active
# bz_list = [25,13,1]
# month = [31,28,31,30,31,30,31,31,30,31,30,31]  # month list
# for i in range(1,4):
#     inday = int(day) - i
#     inday2 = int(day) -i +1;
#     if(inday == 0):
#         remon = int(m) - 1
#         inday = month[remon]
#     elif(inday == -1):
#         remon = int(m) -1
#         inday = month[remon] - 1
#     elif(inday<10):
#         inday = "0{}".format(inday)
#
#     if (inday2 < 1):
#         day = 2
#     elif (inday2 < 10):
#         inday2 = "0{}".format(inday2)
#
#     file = "D:\日报\ss_2021-{}-{}.csv".format(m,inday)
#     # wb = load_workbook(file)
#     # ws = wb.active
#     print(file)
#     csvF = open(file,'r',encoding='utf-8')
#     read = csv.reader(csvF)
#     index = bz_list[i-1]
#     wsmb.cell(index,1).value = "2021/{}/{}".format(m,inday2)
#     index = index +1
#     for line in read:
#         # print(line)
#         k = 1
#         for v in line:
#             wsmb.cell(index, k).value = v
#             k = k+1
#         index = index +1
#
#     # wsmb.cell(index,1).value = "2021/{}/{}".format(m,day)
#     # for j in range(index,index+r):
#     #     for k in range(1,c):
#     #         wsmb.cell(j,k).value = ws.cell(j-index+1,k).value
#
#
# wbmb.save("D:\日报\烽火并发日报2021-{}-{}到{}.xlsx".format(m,inday2,day))
# # 烽火三天日报脚本





# wb = load_workbook("D:\日报\日常巡检表.xlsx")
# ws = wb.active
# print(ws.title)
# print(ws)
# print(ws.max_column)
# print(ws.max_row)
# rmax = ws.max_row
# lmax = ws.max_column
#
# ws.cell(1,lmax+1).value = time.strftime("%y.%m.%d", time.localtime())
# for i in range(2,rmax+1):
#     c = ws.cell(i,lmax).value
#     ws.cell(i,lmax+1).value = c
#     so = ws.cell(i,lmax)
#     tar = ws.cell(i,lmax+1)
#     if so.has_style:   #复杂单元格格式
#         tar._style = copy(so._style)
#         tar.fill = copy(so.fill)
#         tar.border = copy(so.border)
#         tar.alignment = copy(so.alignment)
#
# # ws.merge_cells(start_row=77, start_column=lmax+1, end_row=78, end_column=lmax+1)
# # ws.merge_cells(start_row=79, start_column=lmax+1, end_row=80, end_column=lmax+1)
#
#
# # ws.cell(9,lmax+1).value = 0.52+random.randint(3,6)*0.001+random.randint(0,9)*0.0001
# #
# wb.save("D:\日报\日常巡检表.xlsx")
# # print("修改完成。")
# #
# #
# # #ws.cell(1,lmax+1).value = time.strftime("%m/%d", time.localtime())
# for i in range(1,rmax):
#     c = ws.cell(i,lmax+1).value
#     print(c)