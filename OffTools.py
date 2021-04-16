from selenium import webdriver
from openpyxl import load_workbook
from copy import copy, deepcopy
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import random
from datetime import datetime
from openpyxl import load_workbook
import time
import csv
import os
from ftplib import FTP



def loadDriver():      #加载浏览器驱动
    # driver = webdriver.Edge("H:\software\edgedriver_win64\msedgedriver.exe") #Edge
    driver = webdriver.Chrome("H:\software\chromedriver_win32\chromedriver.exe")  #Chrome
    options = webdriver.ChromeOptions()
    prefs = {"download.default_directory": "D:\日报"}
    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(chrome_options=options)
    time.sleep(5)
    driver.get("http://10.192.70.242:10080/default.aspx")
    driver.implicitly_wait(20)

    hwModel(driver)  # hw日报
    zteModel(driver)  # zte日报
    fhModel(driver)  # fh日报
    time.sleep(300)


# -------------------------------------------------------------------------------------------------

def fhModel(driver):
    print("------------ fh ----------")
    ftb1 = driver.find_element_by_xpath("//*[@id='zz2_QuickLaunchMenun2']/td/table/tbody/tr/td/a")
    # //*[@id="17692"]/tbody/tr/td[1]/a
    ftb1.click()
    ftb2 = driver.find_element_by_xpath("//*[@id='17692']/tbody/tr/td[1]/a")  # 维护作业
    ftb2.click()
    ftb3 = driver.find_element_by_xpath("//*[@id='2283']/tbody/tr/td[1]/a")
    ftb3.click()
    ftb4 = driver.find_element_by_xpath("//*[@id='2314']/tbody/tr/td[1]/a")
    ftb4.click()  # 下载
    ftb5 = driver.find_element_by_xpath("//*[@id='12658']/tbody/tr/td[1]/a")
    ftb5.click()
    time.sleep(20)
    fhUrl = "D:\日报\日常巡检表.xlsx"
    addDataFh(fhUrl)  # 增加数据
    upLoad(driver, fhUrl)  # 上传数据
    time.sleep(3)

def zteModel(driver):
    print("———————— ZTE ——————————")
    # //*[@id="zz2_QuickLaunchMenun2"]/td/table/tbody/tr/td/a
    zbt1 = driver.find_element_by_xpath("//*[@id='zz2_QuickLaunchMenun2']/td/table/tbody/tr/td/a")
    zbt1.click()
    zbt2 = driver.find_element_by_xpath("//*[@id='17692']/tbody/tr/td[1]/a")
    zbt2.click()
    zbt3 = driver.find_element_by_xpath("//*[@id='1']/tbody/tr/td[1]/a")
    zbt3.click()
    zbt6 = driver.find_element_by_xpath("//*[@id='8']/tbody/tr/td[1]/a")
    zbt6.click()
    # //*[@id="17034"]/tbody/tr/td[1]/a
    zbt4 = driver.find_element_by_xpath("//*[@id='17034']/tbody/tr/td[1]/a")  # 不同年份不一样
    zbt4.click()
    # //*[@id="17720"]/tbody/tr/td[1]/a
    zbt5 = driver.find_element_by_xpath("//*[@id='17720']/tbody/tr/td[1]/a")  # 不同月份不一样
    zbt5.click()  # 下载文件
    driver.implicitly_wait(10)
    time.sleep(30)
    zteFileUrl = "D:\日报\\2021ZTE日常维护表_中兴维护计划4月.xlsx"
    addDataZte(zteFileUrl)  # 增加数据
    upLoad(driver, zteFileUrl) #上传数据
    time.sleep(3)

def hwModel(driver):
    print("HW.................")
    bt1 = driver.find_element_by_xpath("//*[@id='zz2_QuickLaunchMenun2']/td/table/tbody/tr/td/a")
    bt1.click()
    driver.implicitly_wait(2)
    bt2 = driver.find_element_by_xpath("//*[@id='17692']/tbody/tr/td[1]/a")
    bt2.click()
    driver.implicitly_wait(2)
    bt3 = driver.find_element_by_xpath("//*[@id='2']/tbody/tr/td[1]/a")
    bt3.click()
    bt6 = driver.find_element_by_xpath("//*[@id='372']/tbody/tr/td[1]/a")
    bt6.click()
    # //*[@id="17032"]/tbody/tr/td[1]/a
    bt4 = driver.find_element_by_xpath("//*[@id='17032']/tbody/tr/td[1]/a")  # 不同年份不一样
    bt4.click()
    # //*[@id="17719"]/tbody/tr/td[1]/a
    bt5 = driver.find_element_by_xpath("//*[@id='17719']/tbody/tr/td[1]/a")  # 不同月份不一样
    bt5.click()  # 下载
    driver.implicitly_wait(10)
    time.sleep(10)
    hwFileUrl = "D:\日报\HW_日常维护表_HW维护计划4月.xlsx"  #文件地址 每月需要修改
    addDataHW(hwFileUrl)   # 处理表格
    upLoad(driver, hwFileUrl) #上传文件
    time.sleep(10)


def upLoad(driver,fileUrl):  #上传文件
    print("开始上传")
    # //*[@id="zz18_UploadMenu"]
    bt6 = driver.find_element_by_xpath("//*[@id='zz18_UploadMenu']")  # 上传
    bt6.click()
    input = driver.find_element_by_xpath("//*[@id='ctl00_PlaceHolderMain_ctl01_ctl02_InputFile']")
    input.send_keys(fileUrl)
    # input.click()
    bt7 = driver.find_element_by_xpath("//*[@id='ctl00_PlaceHolderMain_ctl00_RptControls_btnOK']")
    bt7.click()
    driver.implicitly_wait(10)
    time.sleep(10)
    print("上传成功： "+fileUrl)

def addDataFh(furl):
    print("开始增加数据......FH")
    wb = load_workbook(furl)
    ws = wb.active
    print(ws.title)
    rmax = ws.max_row
    lmax = ws.max_column
    ws.cell(1, lmax + 1).value = time.strftime("%y.%m.%d", time.localtime())
    for i in range(2, rmax + 1):
        c = ws.cell(i, lmax).value
        ws.cell(i, lmax + 1).value = c
        so = ws.cell(i, lmax)
        tar = ws.cell(i, lmax + 1)
        if so.has_style:  # 复杂单元格格式
            tar._style = copy(so._style)
            tar.fill = copy(so.fill)
            tar.border = copy(so.border)
            tar.alignment = copy(so.alignment)
    wb.save(furl)
    print("数据添加完成FH")

def addDataHW(furl):
    print("开始增加数据......HW")
    wb = load_workbook(furl)
    ws = wb.active
    print(ws)
    print(ws.max_column)
    print(ws.max_row)
    rmax = ws.max_row
    lmax = ws.max_column
    so = ws.cell(1, lmax)
    tar = ws.cell(1, lmax + 1)
    weekday = datetime.now().weekday()
    if weekday == 0:
        ws.cell(1, lmax + 1).value = ws.cell(1, lmax).value + 3  # 更新日期 星期一加3
    else:
        ws.cell(1, lmax + 1).value = ws.cell(1, lmax).value + 1  # 更新日期
    if so.has_style:   #复杂单元格格式
        tar._style = copy(so._style)
        tar.fill = copy(so.fill)
    for i in range(2, rmax):  #复制更新数据
        c = ws.cell(i, lmax).value
        ws.cell(i, lmax + 1).value = c
    wb.save(furl)
    print("修改完成。")

def addDataZte(furl):
    print("开始增加数据......ZTE")
    wb = load_workbook(furl)
    ws = wb.active
    print(ws)
    print(ws.max_column)
    print(ws.max_row)
    rmax = ws.max_row
    lmax = ws.max_column

    for i in range(1, rmax):
        c = ws.cell(i, lmax).value
        ws.cell(i, lmax + 1).value = c
        so = ws.cell(i, lmax)
        tar = ws.cell(i, lmax + 1)
        if so.has_style:  # 复杂单元格格式
            tar._style = copy(so._style)
            tar.fill = copy(so.fill)
            tar.border = copy(so.border)
            tar.alignment = copy(so.alignment)

    ws.merge_cells(start_row=77, start_column=lmax + 1, end_row=78, end_column=lmax + 1) #合并单元格
    ws.merge_cells(start_row=79, start_column=lmax + 1, end_row=80, end_column=lmax + 1)

    ws.cell(4, lmax + 1).value = int(time.strftime("%d", time.localtime())) #更新日期
    ws.cell(9, lmax + 1).value = 0.52 + random.randint(3, 6) * 0.001 + random.randint(0, 9) * 0.0001 #更新数据

    wb.save(furl)
    print("修改完成。")


# ftp 获取ZTE过去三天日报
def ftp_getZteFile():
    ftp = FTP()
    # ftp.set_debuglevel(2)
    ftp.connect("124.75.32.237", 21)
    ftp.login("zxin10", "Sos10+Pad6")
    print(ftp.getwelcome())
    ftp.cwd("4ktongji")
    print(ftp.dir())
    localdir = "D:\日报\\"
    if not os.path.exists(localdir):
        os.makedirs(localdir)
    listf = ftp.nlst()
    for file in listf:
        local = os.path.join(localdir, file)
        print("下载", ftp.nlst(file))
        file_handler = open(local, 'wb').write
        ftp.retrbinary('RETR ' + file, file_handler)
    ftp.quit()

# 获取日，用于读取过去三天的日报
def getdate(i,day,m,month):
    inday = int(day) - i
    inday2 = int(day) - i + 1
    if (inday < 1):
        remon = int(m) - 1
        inday = month[remon] + inday
    elif (inday < 10):
        inday = "0{}".format(inday)
    if (inday2 < 1):
        remon = int(m) - 1
        inday2 = month[remon] + inday2
    elif (inday2 < 10):
        inday2 = "0{}".format(inday2)
    return inday,inday2

def getribao(m,day,bz_list,month, wbmb , bz, num):
    # wbmb = load_workbook("D:\日报\烽火并发日报模板.xlsx")
    wsmb = wbmb.active
    for i in range(1, num):
        inday1, inday2 = getdate(i, day, m, month)
        if bz == 0:
            inday1 = inday2
            onedayUrl = "D:\日报\ottnodecpservicestatistics-2021{}{}.csv"
            en = "gbk"
        else:
            onedayUrl = "D:\日报\ss_2021-{}-{}.csv"
            en = "utf-8"
        file = onedayUrl.format(m, inday1)
        print(file)
        csvF = open(file, 'r', encoding= en)
        read = csv.reader(csvF)
        index = bz_list[i - 1]
        wsmb.cell(index, 1).value = "2021/{}/{}".format(m, inday2)
        index = index + 1
        for line in read:
            # print(line)
            k = 1
            for v in line:
                wsmb.cell(index, k).value = v
                k = k + 1
            index = index + 1

def getZteandFhRibao():
    d = input("请输入数字 1：生成一天 2：生成三天")
    num = 4
    if int(d) == 1:
        num = 2
    m = time.strftime("%m", time.localtime())
    day = time.strftime("%d", time.localtime())
    fhbz_list = [25, 13, 1]
    month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # month list
    startday = int(day) - 2

    wbmb = load_workbook("D:\日报\烽火并发日报模板.xlsx")
    getribao(m, day, fhbz_list, month, wbmb, 1, num)  # bz = 1 代表fh日报
    if num == 4:
        wbmb.save("D:\日报\烽火并发日报2021-{}-{}到{}.xlsx".format(m, startday, day))
    else:
        wbmb.save("D:\日报\烽火并发日报2021-{}-{}.xlsx".format(m, day))

    ftp_getZteFile()  # 获取中兴日报文件
    ztebz_list = [17, 9, 1]
    wbmb2 = load_workbook("D:\日报\中兴并发日报模板2.xlsx")
    getribao(m, day, ztebz_list, month, wbmb2, 0, num)  # bz = 0 代表zte日报
    if num == 4:
        wbmb2.save("D:\日报\中兴并发日报2021-{}-{}到{}.xlsx".format(m, startday, day))
    else:
        wbmb2.save("D:\日报\中兴并发日报2021-{}-{}.xlsx".format(m, startday, day))


def getSql():
    iid = input("id：")
    fwqiid = input("服务器id：")
    print("update s830cardinf set terminalid={},serterminalid={} where userid='{}';".format(fwqiid, fwqiid, iid))
    print("select terminalid,serterminalid from s830cardinf where userid='{}';".format(iid))

def main():
    et = '10'
    while et != '0':
        of = input("请输入数字：\n  1，sharepoint日报功能\n  2,4K日报生成\n  3，获取修改EPG服务器SQL\n  0，退出...\n")
        et = of
        if of == '1':
            loadDriver()
        elif of == '2':
            getZteandFhRibao()
        elif of == '3':
            getSql()


if __name__ == '__main__':
   main()